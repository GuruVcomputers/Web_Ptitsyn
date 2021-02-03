from openpyxl.reader.excel import load_workbook


class SaveExcel:
    def __init__(self, src):
        self.wb = load_workbook(src)
        self.ws = self.wb.get_sheet_by_name("Sheet1")
        self.dest = "table1.xlsx"

    # Write the value in the cell defined by row_dest+column_dest
    def write_workbook(self, row_dest, column_dest, value):
        c = self.ws.cell(row=row_dest, column=column_dest)
        c.value = value

    # Save excel file
    def save_excel(self):
        self.wb.save(self.dest)
